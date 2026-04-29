import csv
import re
from datetime import datetime
from io import BytesIO
from io import StringIO
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from ..db import get_db
from ..deps import get_current_user
from ..models import IngestionRun, User

router = APIRouter(prefix="/ingestion", tags=["ingestion"])

ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
ALLOWED_STRATEGIES = {"replace", "append"}


def _normalize_identifier(value: str) -> str:
    normalized = value.strip().lower()
    normalized = re.sub(r"[^a-z0-9_]", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    if not normalized:
        normalized = "coluna"
    if normalized[0].isdigit():
        normalized = f"c_{normalized}"
    return normalized


def _quoted_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _parse_csv(content: bytes, delimiter: str) -> tuple[list[str], list[dict[str, str | None]]]:
    try:
        text_content = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text_content = content.decode("latin-1")

    reader = csv.DictReader(StringIO(text_content), delimiter=delimiter)
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV sem cabeçalho.")
    headers = [h.strip() if h else "" for h in reader.fieldnames]
    rows = list(reader)
    return headers, rows


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, str | None]]]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    first_row = next(rows_iter, None)
    if not first_row:
        raise HTTPException(status_code=400, detail="XLSX vazio.")
    headers = [str(cell).strip() if cell is not None else "" for cell in first_row]
    parsed_rows: list[dict[str, str | None]] = []
    for row in rows_iter:
        row_dict: dict[str, str | None] = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else None
            row_dict[header] = None if value is None else str(value)
        parsed_rows.append(row_dict)
    return headers, parsed_rows


@router.get("/runs")
def list_runs(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    runs = db.query(IngestionRun).order_by(IngestionRun.id.desc()).limit(30).all()
    return [
        {
            "id": run.id,
            "source": run.source,
            "dataset": run.dataset,
            "target_table": run.target_table,
            "strategy": run.strategy,
            "file_name": run.file_name,
            "status": run.status,
            "row_count": run.row_count,
            "created_by_email": run.created_by_email,
            "created_at": run.created_at,
            "finished_at": run.finished_at,
            "error_message": run.error_message,
        }
        for run in runs
    ]


@router.post("/import")
async def import_raw_table(
    file: UploadFile = File(...),
    source: str = Form(...),
    dataset: str = Form(...),
    strategy: str = Form("replace"),
    csv_delimiter: str = Form(";"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    extension = Path(file.filename or "").suffix.lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato inválido. Envie CSV ou XLSX.",
        )
    if strategy not in ALLOWED_STRATEGIES:
        raise HTTPException(status_code=400, detail="Estratégia inválida. Use replace ou append.")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Arquivo vazio.")

    if extension == ".csv":
        headers, rows = _parse_csv(content, csv_delimiter)
    else:
        headers, rows = _parse_xlsx(content)

    normalized_source = _normalize_identifier(source)
    normalized_dataset = _normalize_identifier(dataset)
    target_table = f"{normalized_source}__{normalized_dataset}"

    normalized_map: dict[str, str] = {}
    used_names: set[str] = set()
    for header in headers:
        base = _normalize_identifier(header or "coluna")
        candidate = base
        suffix = 1
        while candidate in used_names:
            suffix += 1
            candidate = f"{base}_{suffix}"
        used_names.add(candidate)
        normalized_map[header] = candidate

    run = IngestionRun(
        source=source,
        dataset=dataset,
        target_table=target_table,
        strategy=strategy,
        file_name=file.filename or "arquivo_sem_nome",
        status="running",
        columns_map=normalized_map,
        created_by_email=current_user.email,
    )
    db.add(run)
    db.commit()
    db.refresh(run)

    try:
        with db.bind.begin() as connection:
            connection.execute(text("CREATE SCHEMA IF NOT EXISTS raw"))
            create_columns_sql = ", ".join(
                f'{_quoted_identifier(col)} TEXT' for col in normalized_map.values()
            )
            connection.execute(
                text(
                    f"CREATE TABLE IF NOT EXISTS raw.{_quoted_identifier(target_table)} "
                    f"(id BIGSERIAL PRIMARY KEY, {create_columns_sql})"
                )
            )

            # If new file has extra columns, add them without dropping old ones.
            existing_columns_result = connection.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = 'raw' AND table_name = :table_name
                    """
                ),
                {"table_name": target_table},
            )
            existing_columns = {row[0] for row in existing_columns_result}
            for normalized_column in normalized_map.values():
                if normalized_column not in existing_columns:
                    connection.execute(
                        text(
                            f"ALTER TABLE raw.{_quoted_identifier(target_table)} "
                            f"ADD COLUMN {_quoted_identifier(normalized_column)} TEXT"
                        )
                    )

            if strategy == "replace":
                connection.execute(text(f"TRUNCATE TABLE raw.{_quoted_identifier(target_table)}"))

            if rows:
                insert_columns = list(normalized_map.values())
                insert_columns_sql = ", ".join(_quoted_identifier(c) for c in insert_columns)
                values_sql = ", ".join(f":{c}" for c in insert_columns)
                insert_stmt = text(
                    f"INSERT INTO raw.{_quoted_identifier(target_table)} ({insert_columns_sql}) "
                    f"VALUES ({values_sql})"
                )

                payload_rows = []
                for row in rows:
                    payload = {}
                    for original_col, normalized_col in normalized_map.items():
                        value = row.get(original_col)
                        payload[normalized_col] = value if value not in ("", None) else None
                    payload_rows.append(payload)

                connection.execute(insert_stmt, payload_rows)

        run.status = "success"
        run.row_count = len(rows)
        run.finished_at = datetime.utcnow()
        db.add(run)
        db.commit()
    except Exception as exc:
        run.status = "failed"
        run.error_message = str(exc)[:500]
        run.finished_at = datetime.utcnow()
        db.add(run)
        db.commit()
        raise HTTPException(status_code=500, detail=f"Falha na ingestão: {str(exc)}") from exc

    return {
        "status": "success",
        "ingestion_run_id": run.id,
        "target_schema": "raw",
        "target_table": target_table,
        "strategy": strategy,
        "row_count": len(rows),
        "columns_count": len(normalized_map),
    }
